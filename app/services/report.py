import pandas as pd
import jdatetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
import calendar
from datetime import datetime, timedelta

SHIFT_TYPES = ["M", "A", "D", "E", "N", "DE"]
HOLIDAYS = [6, 13, 14, 15, 20, 27]

# Define color codes for each shift type (hex)
SHIFT_COLORS = {
    "M": "FFC7CE",  # Light Red
    "A": "FFEB9C",  # Light Yellow
    "D": "C6EFCE",  # Light Green
    "E": "9BC2E6",  # Light Blue
    "N": "D9D2E9",  # Light Purple
    "DE": "F4B084"  # Light Orange
}

def schedule_json_to_excel(schedule_json, excel_filepath):
    """
    Converts the schedule JSON into an Excel file.
    
    schedule_json: List of dictionaries like
      [
        {"M": [0,2], "A": [1], "D": [...], ...},  # Day 0
        {"M": [...], ...},                        # Day 1
        ...
      ]
    excel_filepath: path to save the Excel file (e.g. "schedule.xlsx")
    """

    # Prepare a list to build DataFrame rows
    rows = []

    # For each day, flatten shift-staff info
    for day_idx, day_schedule in enumerate(schedule_json, start=1):
        # For each shift type
        for shift_type, staff_ids in day_schedule.items():
            if not staff_ids:
                # If no staff assigned, put empty row
                rows.append({
                    "Day": day_idx,
                    "Shift": shift_type,
                    "Staff ID": ""
                })
            else:
                for sid in staff_ids:
                    rows.append({
                        "Day": day_idx,
                        "Shift": shift_type,
                        "Staff ID": sid
                    })

    # Create DataFrame
    df = pd.DataFrame(rows, columns=["Day", "Shift", "Staff ID"])

    # Save as Excel
    df.to_excel(excel_filepath, index=False)

    print(f"Excel file saved to {excel_filepath}")







# def create_hospital_style_schedule(schedule_json, excel_filepath, staff_names=None, start_date=None):
#     """
#     Create an Excel file with days as columns, staff as rows,
#     shifts filled in and color-coded.

#     schedule_json: list of dicts per day {shift_type: [staff_ids]}
#     excel_filepath: path to save Excel file
#     staff_names: optional dict mapping staff_id to name
#     start_date: datetime.date or string 'YYYY-MM-DD' for day 1 to calculate weekdays

#     expamle:
#     if sched is not None:
#          create_hospital_style_schedule(
#              sched, # Json
#              "hospital_style_schedule.xlsx",
#              staff_names=None,  # or provide {0: "Alice", 1: "Bob", ...}
#              start_date="2025-07-01"  # adjust this to your month start date
#          )
#     """
#     TOTAL_STAFF = max(max(staff_ids) if staff_ids else 0 for day in schedule_json for staff_ids in day.values()) + 1
    
#     # Default staff_names: just staff_id as string
#     if staff_names is None:
#         staff_names = {i: f"Staff {i}" for i in range(TOTAL_STAFF)}

#     # Prepare DataFrame rows (index=staff, columns=days)
#     columns = []
#     base_date = None
#     if start_date is None:
#         # Default: Assume day 1 is a Monday
#         base_date = datetime.strptime("2025-07-01", "%Y-%m-%d").date()  # adjust to real start date if you want
#     elif isinstance(start_date, str):
#         base_date = datetime.strptime(start_date, "%Y-%m-%d").date()
#     else:
#         base_date = start_date

#     for day_idx in range(len(schedule_json)):
#         current_date = base_date + timedelta(days=day_idx)
#         weekday_name = calendar.day_name[current_date.weekday()]  # e.g. Monday
#         columns.append(f"{day_idx + 1} ({weekday_name[:3]})")

#     # Initialize empty DataFrame with staff as index and days as columns
#     df = pd.DataFrame("", index=[staff_names[i] for i in range(TOTAL_STAFF)], columns=columns)

#     # Fill in shifts
#     for day_idx, day_schedule in enumerate(schedule_json):
#         for shift_type, staff_ids in day_schedule.items():
#             for sid in staff_ids:
#                 current_val = df.at[staff_names[sid], columns[day_idx]]
#                 # If empty, just put shift, else append
#                 if current_val:
#                     df.at[staff_names[sid], columns[day_idx]] = current_val + ", " + shift_type
#                 else:
#                     df.at[staff_names[sid], columns[day_idx]] = shift_type

#     # Save to Excel with color
#     df.to_excel(excel_filepath)

#     # Now open workbook with openpyxl to apply colors and formatting
#     wb = load_workbook(excel_filepath)
#     ws = wb.active

#     # Center align all cells
#     alignment = Alignment(horizontal="center", vertical="center")
    
#     # Map shift types to their color fill
#     fills = {k: PatternFill(start_color=v, end_color=v, fill_type="solid") for k,v in SHIFT_COLORS.items()}

#     # Style header row (days)
#     for col_idx in range(2, 2 + len(columns)):
#         cell = ws.cell(row=1, column=col_idx)
#         cell.alignment = alignment

#     # Style staff name column (first column)
#     for row_idx in range(2, 2 + TOTAL_STAFF):
#         cell = ws.cell(row=row_idx, column=1)
#         cell.alignment = alignment

#     # Apply color fills for shifts
#     for row_idx in range(2, 2 + TOTAL_STAFF):
#         for col_idx in range(2, 2 + len(columns)):
#             cell = ws.cell(row=row_idx, column=col_idx)
#             cell.alignment = alignment
#             cell_value = str(cell.value)
#             if cell_value:
#                 # If multiple shifts, just color with the first shift color
#                 first_shift = cell_value.split(",")[0].strip()
#                 fill = fills.get(first_shift)
#                 if fill:
#                     cell.fill = fill

#     # Adjust column widths
#     for col in ws.columns:
#         max_length = 0
#         col_letter = col[0].column_letter
#         for cell in col:
#             if cell.value:
#                 max_length = max(max_length, len(str(cell.value)))
#         adjusted_width = max_length + 2
#         ws.column_dimensions[col_letter].width = adjusted_width

#     # Adjust row heights (optional)
#     for row_idx in range(1, 2 + TOTAL_STAFF):
#         ws.row_dimensions[row_idx].height = 20

#     wb.save(excel_filepath)
#     print(f"Hospital style Excel schedule saved to {excel_filepath}")


# def create_hospital_style_schedule(schedule_json, excel_filepath, staff_names=None, start_date=None):
#     from openpyxl.styles import PatternFill, Alignment
#     import calendar
#     from datetime import datetime, timedelta
#     import pandas as pd
#     from openpyxl import load_workbook

#     TOTAL_STAFF = max(max(staff_ids) if staff_ids else 0 for day in schedule_json for staff_ids in day.values()) + 1

#     if staff_names is None:
#         staff_names = {i: f"Staff {i}" for i in range(TOTAL_STAFF)}

#     # Prepare DataFrame columns
#     columns = []
#     if start_date is None:
#         base_date = datetime.strptime("2025-07-01", "%Y-%m-%d").date()
#     elif isinstance(start_date, str):
#         base_date = datetime.strptime(start_date, "%Y-%m-%d").date()
#     else:
#         base_date = start_date

#     for day_idx in range(len(schedule_json)):
#         current_date = base_date + timedelta(days=day_idx)
#         weekday_name = calendar.day_name[current_date.weekday()]
#         columns.append(f"{day_idx + 1} ({weekday_name[:3]})")

#     # Initialize empty DataFrame
#     df = pd.DataFrame("", index=[staff_names[i] for i in range(TOTAL_STAFF)], columns=columns)

#     # Fill in shifts
#     for day_idx, day_schedule in enumerate(schedule_json):
#         for shift_type, staff_ids in day_schedule.items():
#             for sid in staff_ids:
#                 current_val = df.at[staff_names[sid], columns[day_idx]]
#                 if current_val:
#                     df.at[staff_names[sid], columns[day_idx]] = current_val + ", " + shift_type
#                 else:
#                     df.at[staff_names[sid], columns[day_idx]] = shift_type

#     # Save to Excel first
#     df.to_excel(excel_filepath)

#     # Open with openpyxl for styling
#     wb = load_workbook(excel_filepath)
#     ws = wb.active

#     alignment = Alignment(horizontal="center", vertical="center")

#     # Shift type color fills
#     fills = {k: PatternFill(start_color=v, end_color=v, fill_type="solid") for k, v in SHIFT_COLORS.items()}

#     # Holiday fill
#     # holiday_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
#     holiday_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")


#     # Style header row (day headers)
#     for col_idx in range(2, 2 + len(columns)):
#         day_number = int(columns[col_idx - 2].split(" ")[0])  # e.g. "6 (Sat)" -> 6
#         header_cell = ws.cell(row=1, column=col_idx)
#         header_cell.alignment = alignment

#         # Color header if it's a holiday
#         if day_number in HOLIDAYS:
#             header_cell.fill = holiday_fill

#     # Style staff name cells
#     for row_idx in range(2, 2 + TOTAL_STAFF):
#         ws.cell(row=row_idx, column=1).alignment = alignment

#     # Fill and color shift cells
#     for row_idx in range(2, 2 + TOTAL_STAFF):
#         for col_idx in range(2, 2 + len(columns)):
#             cell = ws.cell(row=row_idx, column=col_idx)
#             cell.alignment = alignment
#             cell_value = str(cell.value)
#             day_number = int(columns[col_idx - 2].split(" ")[0])

#             if cell_value:
#                 first_shift = cell_value.split(",")[0].strip()
#                 fill = fills.get(first_shift)
#                 if fill:
#                     cell.fill = fill

#             # Apply holiday background behind shift color if it's a holiday (only if empty or after shift color)
#             if day_number in HOLIDAYS and not cell.fill.fgColor.rgb == holiday_fill.fgColor.rgb:
#                 # If already filled with a shift color, no need to override
#                 if cell.fill.patternType is None:  # if no existing fill
#                     cell.fill = holiday_fill

#     # Adjust column widths
#     for col in ws.columns:
#         max_length = 0
#         col_letter = col[0].column_letter
#         for cell in col:
#             if cell.value:
#                 max_length = max(max_length, len(str(cell.value)))
#         ws.column_dimensions[col_letter].width = max_length + 2

#     # Adjust row heights
#     for row_idx in range(1, 2 + TOTAL_STAFF):
#         ws.row_dimensions[row_idx].height = 20

#     wb.save(excel_filepath)
#     print(f"Hospital style Excel schedule saved to {excel_filepath}")



# def create_hospital_style_schedule(schedule_json, excel_filepath, staff_names=None, start_date=None):
#     from openpyxl.styles import PatternFill, Alignment
#     import pandas as pd
#     from openpyxl import load_workbook

#     TOTAL_STAFF = max(max(staff_ids) if staff_ids else 0 for day in schedule_json for staff_ids in day.values()) + 1

#     if staff_names is None:
#         staff_names = {i: f"Staff {i}" for i in range(TOTAL_STAFF)}

#     # Prepare columns — using Persian date if no start_date provided
#     columns = []
#     if start_date is None:
#         # Default to 1403-04-01 (Hijri Shamsi 1 Tir 1403)
#         base_date = jdatetime.date(1403, 4, 1)
#     elif isinstance(start_date, str):
#         # Parse Jalali date from 'YYYY-MM-DD' string
#         y, m, d = map(int, start_date.split("-"))
#         base_date = jdatetime.date(y, m, d)
#     else:
#         base_date = start_date  # Assuming jdatetime.date

#     for day_idx in range(len(schedule_json)):
#         current_date = base_date + jdatetime.timedelta(days=day_idx)
#         weekday_name = current_date.strftime("%A")  # Persian weekday name
#         columns.append(f"{current_date.day} ({weekday_name})")

#     # Initialize empty DataFrame
#     df = pd.DataFrame("", index=[staff_names[i] for i in range(TOTAL_STAFF)], columns=columns)

#     # Fill in shifts
#     for day_idx, day_schedule in enumerate(schedule_json):
#         for shift_type, staff_ids in day_schedule.items():
#             for sid in staff_ids:
#                 current_val = df.at[staff_names[sid], columns[day_idx]]
#                 if current_val:
#                     df.at[staff_names[sid], columns[day_idx]] = current_val + ", " + shift_type
#                 else:
#                     df.at[staff_names[sid], columns[day_idx]] = shift_type

#     # Save to Excel
#     df.to_excel(excel_filepath)

#     # Open with openpyxl to apply styling
#     wb = load_workbook(excel_filepath)
#     ws = wb.active

#     alignment = Alignment(horizontal="center", vertical="center")

#     fills = {k: PatternFill(start_color=v, end_color=v, fill_type="solid") for k, v in SHIFT_COLORS.items()}
#     holiday_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")

#     # Style header row (Persian date headers)
#     for col_idx in range(2, 2 + len(columns)):
#         day_number = int(columns[col_idx - 2].split(" ")[0])
#         header_cell = ws.cell(row=1, column=col_idx)
#         header_cell.alignment = alignment
#         if day_number in HOLIDAYS:
#             header_cell.fill = holiday_fill

#     # Style staff name cells
#     for row_idx in range(2, 2 + TOTAL_STAFF):
#         ws.cell(row=row_idx, column=1).alignment = alignment

#     # Color shifts and apply holiday fill for cells
#     for row_idx in range(2, 2 + TOTAL_STAFF):
#         for col_idx in range(2, 2 + len(columns)):
#             cell = ws.cell(row=row_idx, column=col_idx)
#             cell.alignment = alignment
#             cell_value = str(cell.value)
#             day_number = int(columns[col_idx - 2].split(" ")[0])

#             if cell_value:
#                 first_shift = cell_value.split(",")[0].strip()
#                 fill = fills.get(first_shift)
#                 if fill:
#                     cell.fill = fill

#             if day_number in HOLIDAYS and cell.fill.patternType is None:
#                 cell.fill = holiday_fill

#     # Adjust column widths
#     for col in ws.columns:
#         max_length = 0
#         col_letter = col[0].column_letter
#         for cell in col:
#             if cell.value:
#                 max_length = max(max_length, len(str(cell.value)))
#         ws.column_dimensions[col_letter].width = max_length + 2

#     for row_idx in range(1, 2 + TOTAL_STAFF):
#         ws.row_dimensions[row_idx].height = 20

#     wb.save(excel_filepath)
#     print(f"Hospital style Excel schedule saved to {excel_filepath}")


def create_hospital_style_schedule(schedule_json, excel_filepath, staff_names=None, staff_list=None, start_date=None):
    from openpyxl.styles import PatternFill, Alignment
    import pandas as pd
    from openpyxl import load_workbook

    TOTAL_STAFF = max(max(staff_ids) if staff_ids else 0 for day in schedule_json for staff_ids in day.values()) + 1

    if staff_names is None:
        staff_names = {i: f"Staff {i}" for i in range(TOTAL_STAFF)}

    if staff_list is None:
        staff_list = [None] * TOTAL_STAFF  # Fallback dummy list

    # Prepare columns — using Persian date if no start_date provided
    columns = []
    if start_date is None:
        base_date = jdatetime.date(1403, 4, 1)
    elif isinstance(start_date, str):
        y, m, d = map(int, start_date.split("-"))
        base_date = jdatetime.date(y, m, d)
    else:
        base_date = start_date

    for day_idx in range(len(schedule_json)):
        current_date = base_date + jdatetime.timedelta(days=day_idx)
        weekday_name = current_date.strftime("%A")
        columns.append(f"{current_date.day} ({weekday_name})")

    # Prepare initial DataFrame with profession and gender columns
    data = []
    for i in range(TOTAL_STAFF):
        staff_name = staff_names.get(i, f"Staff {i}")
        role = staff_list[i].role.value if staff_list[i] else "Unknown"
        gender = staff_list[i].gender if staff_list[i] else "Unknown"
        sup_type = staff_list[i].supervisor_type if staff_list[i] else "None"
        row = {"Name": staff_name, "Supervisor Type": sup_type, "Role": role, "Gender": gender}
        data.append(row)

    df = pd.DataFrame(data)
    for col in columns:
        df[col] = ""

    # Fill in shifts
    for day_idx, day_schedule in enumerate(schedule_json):
        for shift_type, staff_ids in day_schedule.items():
            for sid in staff_ids:
                current_val = df.at[sid, columns[day_idx]]
                if current_val:
                    df.at[sid, columns[day_idx]] = current_val + ", " + shift_type
                else:
                    df.at[sid, columns[day_idx]] = shift_type

    # Save to Excel
    df.to_excel(excel_filepath, index=False)

    # Open with openpyxl for styling
    wb = load_workbook(excel_filepath)
    ws = wb.active

    alignment = Alignment(horizontal="center", vertical="center")
    fills = {k: PatternFill(start_color=v, end_color=v, fill_type="solid") for k, v in SHIFT_COLORS.items()}
    # holiday_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid") # light red
    holiday_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid") # powerfull red

    num_info_columns = 4  # Name, Role, Gender , supervisor type

    # Style header cells (Persian date headers)
    for col_idx in range(num_info_columns + 1, num_info_columns + 1 + len(columns)):
        day_number = int(columns[col_idx - num_info_columns - 1].split(" ")[0])
        header_cell = ws.cell(row=1, column=col_idx)
        header_cell.alignment = alignment
        if day_number in HOLIDAYS:
            header_cell.fill = holiday_fill

    # Style data cells
    for row_idx in range(2, 2 + TOTAL_STAFF):
        for col_idx in range(1, num_info_columns + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = alignment

        for col_idx in range(num_info_columns + 1, num_info_columns + 1 + len(columns)):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = alignment
            cell_value = str(cell.value)
            day_number = int(columns[col_idx - num_info_columns - 1].split(" ")[0])

            if cell_value:
                first_shift = cell_value.split(",")[0].strip()
                fill = fills.get(first_shift)
                if fill:
                    cell.fill = fill

            if day_number in HOLIDAYS and cell.fill.patternType is None:
                cell.fill = holiday_fill

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    for row_idx in range(1, 2 + TOTAL_STAFF):
        ws.row_dimensions[row_idx].height = 20

    wb.save(excel_filepath)
    print(f"Hospital style Excel schedule saved to {excel_filepath}")
